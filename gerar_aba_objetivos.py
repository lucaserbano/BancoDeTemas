from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/Users/lucas/Downloads/Paôla - Calendário Editorial 2026/Calendário Editorial 2026 — Medicina do Estilo de Vida.xlsx"

wb = load_workbook(XLSX)

# Remove aba se já existir (re-execução)
if "Objetivos Mês a Mês" in wb.sheetnames:
    del wb["Objetivos Mês a Mês"]

ws = wb.create_sheet("Objetivos Mês a Mês", 0)  # primeira aba

# ── PALETA ────────────────────────────────────────────────────────────────────
HEADER_BG    = "1A3A4A"
SUBHEADER_BG = "2E6B8A"
SEPT_BG      = "FFF3CD"   # destaque Setembro
ROW_ALT      = "F4F9FC"
ROW_BASE     = "FFFFFF"

LINHA_CORES = {
    "Medicina do Estilo de Vida": ("D6EAF8", "1A5276"),
    "Conversão":                  ("FDEDEC", "922B21"),
    "Causa & Conscientização":    ("D5F5E3", "1E8449"),
    "Humanização":                ("F4ECF7", "76448A"),
}

FUNIL_CORES = {
    "TOPO":  ("D6EAF8", "1A5276"),
    "MEIO":  ("FDEBD0", "784212"),
    "FUNDO": ("FDEDEC", "922B21"),
}

# ── DADOS ─────────────────────────────────────────────────────────────────────
# Colunas: Mês | # | Objetivo do mês | Posts planejados | Linhas editoriais | Datas importantes | Funil dominante
meses = [
    {
        "mes":      "Abril",
        "num":      "1º mês",
        "objetivo": (
            "Lançamento do perfil renovado + Prevenção cardiovascular\n"
            "Primeiro mês de publicação: estreia com apresentação em SP (reposicionamento que estava previsto para janeiro) "
            "ancorada no Dia Mundial da Saúde (7/abr). Estabelecer voz e identidade do perfil."
        ),
        "posts": (
            "1. Apresentação — nova fase em SP (Conv)\n"
            "2. Checkup: quem precisa fazer? (MEV)\n"
            "3. MAPA: quando a pressão no consultório engana (MEV)\n"
            "4. Meus diferenciais como especialista (Conv)"
        ),
        "linhas":  "MEV (2) · Conv (2)",
        "datas":   "7/abr — Dia Mundial da Saúde",
        "funil":   "TOPO + FUNDO",
        "destaque": False,
    },
    {
        "mes":      "Maio",
        "num":      "2º mês",
        "objetivo": (
            "Saúde cardiovascular da mulher + Dia das Mães\n"
            "Usar o gancho emocional do Dia das Mães (10/mai) para conectar cuidado com a saúde do coração. "
            "Incorporar temas de estresse e lifestyle que estavam planejados para fevereiro."
        ),
        "posts": (
            "1. Saúde cardiovascular da mulher (MEV)\n"
            "2. Arritmias e IC — sintomas a não ignorar (MEV)\n"
            "3. Cuidar da saúde = presente para a família (Conv)\n"
            "4. Congresso ou corrida com gancho cardiovascular (Hum) — se houver evento"
        ),
        "linhas":  "MEV (2) · Conv (1) · Hum (1)",
        "datas":   "10/mai — Dia das Mães",
        "funil":   "TOPO + FUNDO",
        "destaque": False,
    },
    {
        "mes":      "Junho",
        "num":      "3º mês",
        "objetivo": (
            "Exercício físico e cardiologia\n"
            "Posicionar exercício como prescrição médica cardiovascular. "
            "Gancho do inverno (retomada de atividade física). Apresentar VO2max e qualidade de vida pós-transplante."
        ),
        "posts": (
            "1. Por que cardiologista prescreve exercício? (MEV)\n"
            "2. VO2max: o número que prevê sua longevidade (MEV)\n"
            "3. Qualidade de vida pós-transplante (Causa)\n"
            "4. Como agendar sua consulta (Conv)"
        ),
        "linhas":  "MEV (2) · Causa (1) · Conv (1)",
        "datas":   "—",
        "funil":   "TOPO + MEIO + FUNDO",
        "destaque": False,
    },
    {
        "mes":      "Julho",
        "num":      "4º mês",
        "objetivo": (
            "Posicionamento profissional + IC avançada\n"
            "Consolidar autoridade na especialidade de IC avançada e transplante cardíaco. "
            "Aproveitar eventual congresso para humanização com gancho médico."
        ),
        "posts": (
            "1. IC avançada: o Quarteto Fantástico (MEV)\n"
            "2. Congresso / bastidores do ambulatório (Hum) — se houver\n"
            "3. Por que consultar especialista em IC? (Conv)\n"
            "4. Colesterol e aterosclerose (MEV)"
        ),
        "linhas":  "MEV (2) · Hum (1) · Conv (1)",
        "datas":   "Verificar: SOCESP / CBC",
        "funil":   "TOPO + MEIO + FUNDO",
        "destaque": False,
    },
    {
        "mes":      "Agosto",
        "num":      "5º mês",
        "objetivo": (
            "Fatores de risco cardiovascular + aquecimento para Setembro\n"
            "Resgatar temas de hipertensão, diabetes, sono e dieta DASH que estavam planejados para fev/mar "
            "e não foram publicados. Preparar audiência para a grande campanha de setembro."
        ),
        "posts": (
            "1. Dieta anti-inflamatória / DASH (MEV)\n"
            "2. Sono e o coração (MEV)\n"
            "3. Fila de espera pelo transplante (Causa)\n"
            "4. ECG: essa folha diz muito sobre você (MEV)"
        ),
        "linhas":  "MEV (3) · Causa (1)",
        "datas":   "—",
        "funil":   "TOPO + MEIO",
        "destaque": False,
    },
    {
        "mes":      "Setembro ⭐",
        "num":      "6º mês",
        "objetivo": (
            "CAMPANHA: Mês do Coração — maior mês do ano\n"
            "Concentração máxima de causa e conscientização. Pico de alcance orgânico. "
            "5 posts (bônus de engajamento). Ancoragem nas datas mais importantes do calendário cardiovascular."
        ),
        "posts": (
            "1. Abertura Mês do Coração — dados de mortalidade (Causa)\n"
            "2. Como preservar o coração: guia prático (MEV)\n"
            "3. 27/set — Dia Nacional de Doação de Órgãos (Causa)\n"
            "4. 29/set — Dia Mundial do Coração (Conv)\n"
            "BÔNUS: 'Você é doador?' — post de engajamento (Causa)"
        ),
        "linhas":  "Causa (3) · MEV (1) · Conv (1)",
        "datas":   "27/set — Dia Doação de Órgãos\n29/set — Dia Mundial do Coração",
        "funil":   "TOPO + FUNDO",
        "destaque": True,
    },
    {
        "mes":      "Outubro",
        "num":      "7º mês",
        "objetivo": (
            "Reconversão pós-campanha\n"
            "Aproveitar o aquecimento gerado em setembro para converter novos seguidores em pacientes. "
            "Aprofundar temas técnicos (miocardiopatias, medicamentos) e abrir espaço para Q&A."
        ),
        "posts": (
            "1. Miocardiopatias: quando o músculo do coração adoece (MEV)\n"
            "2. Medicamentos cardiológicos: dúvidas frequentes (MEV)\n"
            "3. Resultados do transplante: dados reais (Causa)\n"
            "4. Q&A com seguidores (Conv)"
        ),
        "linhas":  "MEV (2) · Causa (1) · Conv (1)",
        "datas":   "—",
        "funil":   "MEIO + FUNDO",
        "destaque": False,
    },
    {
        "mes":      "Novembro",
        "num":      "8º mês",
        "objetivo": (
            "Prevenção — saúde preventiva\n"
            "Mês de reforço preventivo. Tabagismo (realocado de março). "
            "Humanização com aniversário ou congresso se houver. Reafirmar diferenciais antes do dezembro."
        ),
        "posts": (
            "1. Tabagismo e o coração (MEV)\n"
            "2. Teste ergométrico: por que o médico pediu a esteira? (MEV)\n"
            "3. Aniversário da Paôla ou congresso (Hum) — se aplicável\n"
            "4. Por que me escolher como sua cardiologista (Conv)"
        ),
        "linhas":  "MEV (2) · Hum (1) · Conv (1)",
        "datas":   "Verificar: Congresso Medicina do Estilo de Vida",
        "funil":   "TOPO + FUNDO",
        "destaque": False,
    },
    {
        "mes":      "Dezembro",
        "num":      "9º mês",
        "objetivo": (
            "Fechamento do ano + cuidados nas festas\n"
            "Encerrar 2026 com conteúdo prático para as festas, humanização (retrospectiva) e CTA de agenda para 2027."
        ),
        "posts": (
            "1. Coração nas festas: álcool, sal e estresse (MEV)\n"
            "2. Coração artificial como ponte para o transplante (Causa)\n"
            "3. Retrospectiva 2026 (Hum)\n"
            "4. Agende seu checkup para 2027 (Conv)"
        ),
        "linhas":  "MEV (1) · Causa (1) · Hum (1) · Conv (1)",
        "datas":   "Festas de fim de ano",
        "funil":   "TOPO + FUNDO",
        "destaque": False,
    },
]

# ── HELPERS ───────────────────────────────────────────────────────────────────
thin  = Side(style="thin",   color="CCCCCC")
thick = Side(style="medium", color="888888")
border_light = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color, bold=False, size=10, italic=False):
    return Font(name="Calibri", color=hex_color, bold=bold, size=size, italic=italic)

def align(h="left", wrap=True, v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── DIMENSÕES ─────────────────────────────────────────────────────────────────
# Colunas: A=Mês, B=#, C=Objetivo, D=Posts planejados, E=Linhas editoriais, F=Datas, G=Funil
col_widths = [18, 11, 60, 52, 22, 30, 18]
headers    = ["MÊS", "SEQUÊNCIA", "OBJETIVO DO MÊS", "POSTS PLANEJADOS", "LINHAS EDITORIAIS", "DATAS IMPORTANTES", "FUNIL DOMINANTE"]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── TÍTULO ────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "CALENDÁRIO EDITORIAL 2026 — OBJETIVOS MÊS A MÊS"
c.fill = fill(HEADER_BG)
c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
c.alignment = align("center", v="center")
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = (
    "@paolapreto · Publicações a partir de ABRIL 2026 · "
    "9 meses ativos (abr–dez) · Temas de jan–mar redistribuídos em agosto e novembro"
)
c.fill = fill(SUBHEADER_BG)
c.font = Font(name="Calibri", color="FFFFFF", size=10, italic=True)
c.alignment = align("center", v="center")
ws.row_dimensions[2].height = 20

# ── CABEÇALHOS ────────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 30
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=ci, value=h)
    c.fill = fill(SUBHEADER_BG)
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    c.alignment = align("center", wrap=True, v="center")
    c.border = Border(
        left=Side(style="medium", color="FFFFFF"),
        right=Side(style="medium", color="FFFFFF"),
        bottom=Side(style="medium", color="FFFFFF"),
    )

# ── LINHAS DE DADOS ───────────────────────────────────────────────────────────
for ri, m in enumerate(meses, start=4):
    ws.row_dimensions[ri].height = 110

    is_sept = m["destaque"]
    row_bg  = SEPT_BG if is_sept else (ROW_ALT if ri % 2 == 0 else ROW_BASE)

    values = [
        m["mes"],
        m["num"],
        m["objetivo"],
        m["posts"],
        m["linhas"],
        m["datas"],
        m["funil"],
    ]

    for ci, val in enumerate(values, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = fill(row_bg)
        c.border = border_light

        if ci == 1:  # Mês
            c.font  = Font(name="Calibri", color="1A3A4A" if not is_sept else "856404", bold=True, size=12)
            c.alignment = align("center", v="center")
        elif ci == 2:  # Sequência
            c.font      = Font(name="Calibri", color="888888", size=9, italic=True)
            c.alignment = align("center", v="center")
        elif ci == 3:  # Objetivo (primeira linha em negrito)
            c.font      = font("2C3E50", size=10)
            c.alignment = align("left", v="top")
        elif ci == 4:  # Posts
            c.font      = font("4A4A4A", size=9)
            c.alignment = align("left", v="top")
        elif ci == 5:  # Linhas editoriais
            c.font      = font("1A5276", bold=True, size=9)
            c.alignment = align("center", v="center")
        elif ci == 6:  # Datas
            c.font      = font("922B21" if m["datas"] != "—" else "AAAAAA", bold=(m["datas"] != "—"), size=9)
            c.alignment = align("center", v="center")
        elif ci == 7:  # Funil
            c.font      = font("1A3A4A", bold=True, size=9)
            c.alignment = align("center", v="center")

    # borda superior mais grossa entre meses
    for ci in range(1, 8):
        c = ws.cell(row=ri, column=ci)
        c.border = Border(
            left=c.border.left, right=c.border.right, bottom=c.border.bottom,
            top=thick,
        )

# ── RODAPÉ / LEGENDA ──────────────────────────────────────────────────────────
last = len(meses) + 4 + 1
ws.merge_cells(f"A{last}:G{last}")
c = ws[f"A{last}"]
c.value = (
    "Legenda: MEV = Medicina do Estilo de Vida · Conv = Conversão · Causa = Causa & Conscientização · Hum = Humanização  |  "
    "⭐ Setembro = mês prioritário com até 5 posts  |  Temas de jan–mar (hipertensão, diabetes, tabagismo, estresse) redistribuídos em ago/nov"
)
c.fill  = fill("EAF4FB")
c.font  = Font(name="Calibri", color="2E6B8A", size=9, italic=True)
c.alignment = align("center", wrap=True, v="center")
ws.row_dimensions[last].height = 30

# ── CONGELAR CABEÇALHO ────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── SALVAR ────────────────────────────────────────────────────────────────────
wb.save(XLSX)
print(f"Aba 'Objetivos Mês a Mês' adicionada em: {XLSX}")
