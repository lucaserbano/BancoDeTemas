from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Medicina do Estilo de Vida"

# ── PALETA DE CORES ───────────────────────────────────────────────────────────
HEADER_BG = "1A3A4A"

SUBTEMA_CORES = {
    "Exercício Físico":         ("D6EAF8", "1A5276"),
    "Nutrição & Dieta":         ("D5F5E3", "1E8449"),
    "Fatores de Risco":         ("FDEDEC", "922B21"),
    "Sono & Estresse":          ("F4ECF7", "76448A"),
    "Exames & Diagnóstico":     ("FEF9E7", "9A7D0A"),
    "Doenças Cardiovasculares": ("FDEBD0", "A04000"),
}

FUNIL_CORES = {
    "TOPO":  ("D6EAF8", "1A5276"),
    "MEIO":  ("FDEBD0", "784212"),
    "FUNDO": ("FDEDEC", "922B21"),
}

FORMATO_CORES = {
    "Reels":      ("E8DAEF", "6C3483"),
    "Carrossel":  ("D6EAF8", "1A5276"),
    "Post único": ("D5F5E3", "1E8449"),
}

SERIE_CORES = {
    "Mito ou Verdade? 🫀":       ("FFF3CD", "856404"),
    "Pergunte à Dra. Paôla 🩺": ("D1ECF1", "0C5460"),
    "—":                         ("F8F9FA", "6C757D"),
}

# ── DADOS ─────────────────────────────────────────────────────────────────────
# Colunas: #, SUBTEMA, TEMA/TÍTULO, FORMATO, FUNIL,
#          OBJETIVO, GANCHO DE ABERTURA, ORIENTAÇÕES DE PRODUÇÃO, SÉRIE, MÊS

topics = [
    # ── EXERCÍCIO FÍSICO (8 temas) ────────────────────────────────────────────
    (1, "Exercício Físico",
     "Exercício físico como prescrição médica",
     "Reels", "TOPO",
     "Posicionar exercício como remédio cardiovascular comprovado — não como sugestão de estilo de vida",
     "Exercício não é sugestão de estilo de vida. É prescrição médica. E eu vou te explicar por quê.",
     "Reels ≤60s. Abertura com afirmação provocativa. Dado: pacientes que se exercitam têm X% menos risco de morte cardiovascular. Linguagem de amiga médica, sem jargão. CTA fundo de funil: 'Agende sua avaliação para receber a sua prescrição.'",
     "—", "Janeiro"),

    (2, "Exercício Físico",
     "Frequência cardíaca ideal para treinar",
     "Carrossel", "MEIO",
     "Ensinar pacientes a monitorar a intensidade do exercício com segurança e sem medo",
     "Você está se exercitando forte demais — ou fraco demais. Como saber qual é o seu caso?",
     "Carrossel 6 slides. S1: pergunta-gancho. S2: o que é FC máxima — fórmula em linguagem leiga ('220 menos sua idade'). S3–4: zonas de treino com exemplos do dia a dia ('respiração ofegante = zona 3'). S5: sinais de alerta. S6: CTA 'Salva esse post antes do próximo treino.'",
     "—", "Junho"),

    (3, "Exercício Físico",
     "VO2max: o número que prevê sua longevidade",
     "Reels", "MEIO",
     "Apresentar VO2max como marcador de saúde cardiovascular acessível e melhorável",
     "Tem um número que prevê quanto tempo você vai viver. Seu cardiologista já te perguntou sobre ele?",
     "Reels ≤60s. Abertura com dado surpresa. Explicar VO2max como 'capacidade do coração de entregar oxigênio ao corpo'. Citar estudo de mortalidade. Boa notícia: melhora com treino regular. CTA: 'Deixa nos comentários se você conhecia esse número.'",
     "—", "Junho"),

    (4, "Exercício Físico",
     "Cardio vs musculação: qual é melhor para o coração?",
     "Carrossel", "TOPO",
     "Desmistificar que apenas aeróbico protege o coração e ampliar a percepção de exercício saudável",
     "Você vai à academia mas só faz cardio? Seu coração precisa mais do que isso.",
     "Carrossel 6 slides. S1: gancho. S2: benefícios do aeróbico para o coração. S3: benefícios surpreendentes da musculação (reduz PA, melhora sensibilidade à insulina). S4: combinação ideal. S5: cuidados para quem tem doença cardíaca. S6: CTA 'Compartilha com quem só faz uma coisa.'",
     "—", "Maio"),

    (5, "Exercício Físico",
     "Quanto exercício por semana protege o coração?",
     "Post único", "TOPO",
     "Dar uma meta simples e acionável para o público sedentário começar a se mover",
     "30 minutos por dia, 5 vezes por semana. É tudo que o seu coração pede.",
     "Post único com infográfico limpo: pictograma 5 dias × 30 min. Legenda: desmembrar os 150 min semanais em pequenas metas ('pode ser 3x de 10 min'). Fácil de salvar e compartilhar. CTA: 'Compartilha com quem precisa começar hoje.'",
     "—", "Janeiro"),

    (6, "Exercício Físico",
     "Posso fazer exercício se tenho pressão alta?",
     "Reels", "MEIO",
     "Quebrar o medo de que hipertensos não podem se exercitar — crença que piora o quadro",
     "Dra., tenho pressão alta. Posso malhar? Essa dúvida pode estar te impedindo de proteger o coração.",
     "Reels ≤60s. Abrir com a voz do paciente. Resposta direta: sim, com acompanhamento. Explicar que exercício REDUZ a pressão a longo prazo. Quando NÃO fazer (PA muito elevada sem controle). CTA fundo de funil: 'Agende sua avaliação — prescrição individualizada.'",
     "Pergunte à Dra. Paôla 🩺", "Março"),

    (7, "Exercício Físico",
     "Exercício e insuficiência cardíaca: é seguro?",
     "Reels", "MEIO",
     "Mostrar que pacientes com IC podem e devem se exercitar — com supervisão",
     "Meu paciente me disse: 'Prefiro não arriscar.' Mas parar de se mover com insuficiência cardíaca pode piorar tudo.",
     "Reels ≤60s. Abrir com a crença errada do paciente. Virada: reabilitação cardíaca reduz hospitalização. Dado de impacto. Quem indica e como funciona a prescrição segura. CTA: 'Encaminha pra quem você conhece com IC.'",
     "—", "Outubro"),

    (8, "Exercício Físico",
     "Caminhar conta como exercício para o coração?",
     "Post único", "TOPO",
     "Incentivar o público sedentário a começar com algo acessível — remover barreiras de entrada",
     "Você não precisa de academia. Nem de tênis caro. Nem de tempo sobrando. Só precisa começar.",
     "Post único com visual motivacional. Legenda: caminhada de 30 min reduz risco de infarto em X%. Meta simples: 'comece com 15 minutos hoje'. Tom acolhedor, sem pressão. CTA: 'Salva esse post pra lembrar amanhã de manhã.'",
     "—", "Abril"),

    # ── NUTRIÇÃO & DIETA (7 temas) ────────────────────────────────────────────
    (9, "Nutrição & Dieta",
     "Dieta DASH: a mais recomendada por cardiologistas",
     "Carrossel", "TOPO",
     "Apresentar a dieta DASH como estratégia comprovada de controle pressórico em linguagem acessível",
     "Essa dieta reduz a pressão arterial em semanas. E a maioria das pessoas nunca ouviu falar dela.",
     "Carrossel 7 slides. S1: dado de impacto sobre pressão alta. S2: o que é a DASH em 1 frase. S3–6: componentes práticos (frutas, vegetais, laticínios magros, redução de sal) com exemplos de alimentos do dia a dia. S7: CTA 'Salva e começa essa semana.'",
     "—", "Agosto"),

    (10, "Nutrição & Dieta",
     "Alimentos que inflamam o coração (e os substitutos)",
     "Carrossel", "TOPO",
     "Mostrar quais alimentos aumentam inflamação cardiovascular e dar opções concretas de troca",
     "Você come sem saber que está inflamando o seu coração. Vou te mostrar o que trocar.",
     "Carrossel estilo 'vilão vs herói': cada slide = 1 alimento problemático + 1 substituto saudável com foto. Tom prático, sem julgamento. Fácil de salvar. CTA: 'Qual troca você consegue fazer essa semana? Comenta aqui.'",
     "—", "Agosto"),

    (11, "Nutrição & Dieta",
     "Gordura faz mal ao coração? Mito ou verdade",
     "Reels", "TOPO",
     "Desmistificar a visão de que toda gordura é ruim — diferenciar saturada de insaturada",
     "Mito ou Verdade: gordura faz mal ao coração. Você vai se surpreender com a resposta.",
     "Reels ≤60s, formato Mito ou Verdade. Abertura: 'Verdade — mas só metade da história.' Distinguir gordura saturada (vilã) de insaturada (aliada) com exemplos concretos: manteiga vs azeite, embutido vs abacate. CTA: 'Compartilha com quem ainda tem medo de gordura.'",
     "Mito ou Verdade? 🫀", "Março"),

    (12, "Nutrição & Dieta",
     "Sal e pressão alta: quanto é seguro por dia?",
     "Post único", "TOPO",
     "Dar referência numérica clara sobre sódio e gerar consciência sobre consumo excessivo",
     "Você provavelmente está consumindo o dobro do sal recomendado — e não sabe.",
     "Post único com infográfico: colher de sal visual + número seguro (5g/dia OMS). Legenda: 1 dica prática para reduzir sal sem sofrimento ('espere 10 min antes de salgar'). Tom direto, sem alarme. CTA: 'Salva esse post e mede o seu sal hoje.'",
     "—", "Março"),

    (13, "Nutrição & Dieta",
     "Suplementos com evidência real para o coração",
     "Carrossel", "MEIO",
     "Orientar sobre suplementos com evidência científica e desmistificar modas sem base",
     "Ômega-3, coenzima Q10, vitamina D... Quais realmente funcionam? A resposta não é simples.",
     "Carrossel 6 slides. S1: gancho com lista de suplementos populares. S2–5: cada um com veredito ('evidência forte / fraca / sem evidência') em linguagem leiga. S6: 'nenhum substitui medicamento prescrito' + CTA: 'Pergunta pro seu cardiologista antes de comprar qualquer coisa.'",
     "Mito ou Verdade? 🫀", "Outubro"),

    (14, "Nutrição & Dieta",
     "O que comer antes e depois do treino para o coração",
     "Carrossel", "TOPO",
     "Cruzar nutrição e exercício para engajar público ativo e mostrar cuidado integrado",
     "Você treina mas não sabe o que comer? Seu coração sente a diferença.",
     "Carrossel 5 slides. S1: gancho. S2: antes do treino (carboidrato leve, hidratação). S3: depois do treino (proteína + hidratação). S4: o que EVITAR (cafeína em excesso, jejum prolongado). S5: CTA 'Salva e leva pro próximo treino.'",
     "—", "Junho"),

    (15, "Nutrição & Dieta",
     "Álcool e coração: existe dose segura?",
     "Reels", "TOPO",
     "Atualizar o público sobre as evidências mais recentes que revisaram o 'vinho faz bem'",
     "Dra., uma taça de vinho por dia faz bem para o coração? A ciência mudou de ideia — e eu preciso te contar.",
     "Reels ≤60s. Abrir com a voz do paciente. Tom: 'eu sei que você não quer ouvir isso'. Citar revisão das diretrizes mais recentes. Sem julgamento, sem pregação. CTA: 'Compartilha com quem sempre usou essa justificativa.'",
     "Mito ou Verdade? 🫀", "Dezembro"),

    # ── FATORES DE RISCO (8 temas) ────────────────────────────────────────────
    (16, "Fatores de Risco",
     "Hipertensão: a doença que não avisa",
     "Reels", "TOPO",
     "Criar consciência sobre a prevalência da hipertensão silenciosa e motivar check-up",
     "Pressão alta não dói. Não avisa. E quando o coração finalmente sente — já passou tempo demais.",
     "Reels ≤60s. Abertura emocional e direta. Dado: 1 em 3 brasileiros tem hipertensão sem saber. Sequência: 'como descobrir' → 'o que fazer'. CTA fundo de funil forte: 'Você sabe sua pressão hoje? Se não, agenda uma consulta.'",
     "—", "Janeiro"),

    (17, "Fatores de Risco",
     "Diabetes tipo 2 e o risco duplo para o coração",
     "Carrossel", "MEIO",
     "Alertar diabéticos e pré-diabéticos sobre o risco cardiovascular aumentado e a importância do acompanhamento",
     "Se você tem diabetes, o risco para o seu coração já é o dobro. Você está cuidando dos dois?",
     "Carrossel 6 slides. S1: dado de impacto. S2: por que diabetes danifica os vasos (linguagem simples: 'açúcar desgasta as paredes'). S3: exames para monitorar os dois juntos. S4: metas de controle. S5: o papel do cardiologista nessa história. S6: CTA 'Agende uma avaliação cardiovascular.'",
     "—", "Março"),

    (18, "Fatores de Risco",
     "Colesterol alto: remédio ou só dieta?",
     "Reels", "MEIO",
     "Orientar sobre quando a estatina é necessária e desmistificar o medo do medicamento",
     "Dra., meu colesterol tá alto. Preciso mesmo tomar remédio? Essa dúvida chega toda semana no meu consultório.",
     "Reels ≤60s. Abrir com a voz do paciente. Resposta direta: depende do risco global, não só do número. Explicar 'risco cardiovascular' em 1 frase simples. Quando só dieta basta vs quando remédio é necessário. CTA: 'Deixa nos comentários o que você já ouviu sobre estatina.'",
     "Pergunte à Dra. Paôla 🩺", "Outubro"),

    (19, "Fatores de Risco",
     "Tabagismo: o que acontece com o coração a cada cigarro",
     "Carrossel", "TOPO",
     "Mostrar o impacto imediato e progressivo do cigarro no sistema cardiovascular",
     "Você sabe o que acontece com o seu coração nos 20 minutos depois que você acende um cigarro?",
     "Carrossel cronológico de impacto. S1: gancho com pergunta. S2: '2 min — FC e PA sobem'. S3: '20 min — vasos ainda contraídos'. S4: '1 ano sem fumar — risco de infarto cai à metade'. S5: onde buscar ajuda. S6: CTA 'Compartilha com quem quer largar o cigarro.'",
     "—", "Novembro"),

    (20, "Fatores de Risco",
     "Gordura abdominal e risco cardiovascular",
     "Post único", "TOPO",
     "Associar gordura visceral ao risco cardíaco sem julgamento estético — foco em saúde",
     "Não é sobre estética. É sobre o coração.",
     "Post único com visual limpo e afirmação direta. Legenda: circunferência abdominal de risco (homens >94cm, mulheres >80cm) + por que essa gordura é diferente. Tom acolhedor, zero julgamento. CTA: 'Salva esse post e mede hoje em casa.'",
     "—", "Abril"),

    (21, "Fatores de Risco",
     "Sedentarismo: o risco invisível do dia a dia",
     "Reels", "TOPO",
     "Criar urgência sobre o sedentarismo para motivar pequenas mudanças de comportamento",
     "Ficar sentado por 8 horas por dia é tão perigoso para o coração quanto fumar. Você sabia disso?",
     "Reels ≤60s. Dado de impacto na abertura (deixar a afirmação respirar). Explicar mecanismo: inflamação crônica, PA elevada, resistência à insulina. Dicas práticas de como quebrar o sedentarismo sem academia. CTA: 'Levanta e faz 5 minutos de caminhada agora.'",
     "—", "Janeiro"),

    (22, "Fatores de Risco",
     "Histórico familiar de doença cardíaca: o que fazer?",
     "Carrossel", "TOPO",
     "Sensibilizar quem tem parentes com doença cardíaca a fazer rastreamento precoce",
     "Seu pai ou mãe teve infarto? Isso importa — mas não é destino. Vou te explicar.",
     "Carrossel 5 slides. S1: gancho empoderador. S2: o que herança genética significa de verdade. S3: fatores que você PODE controlar. S4: quais exames pedir e quando. S5: CTA 'Conta nos comentários se tem histórico familiar.'",
     "—", "Fevereiro"),

    (23, "Fatores de Risco",
     "Estresse crônico como fator de risco cardiovascular",
     "Reels", "TOPO",
     "Conectar saúde mental à saúde cardiovascular — ampliar a percepção de prevenção",
     "O estresse que você normaliza pode estar matando o seu coração. E isso não é figura de linguagem.",
     "Reels ≤60s. Abertura emocional forte. Mecanismo em linguagem simples: cortisol → inflamação → artérias. Conexão com sono, exercício e IC. Tom: 'você não está exagerando, seu corpo sente de verdade'. CTA: 'Compartilha com quem precisa ouvir isso.'",
     "—", "Fevereiro"),

    # ── SONO & ESTRESSE (4 temas) ─────────────────────────────────────────────
    (24, "Sono & Estresse",
     "Apneia do sono e o coração: a conexão que poucos conhecem",
     "Reels", "MEIO",
     "Alertar sobre apneia obstrutiva como fator de risco cardiovascular subestimado",
     "Você ronca? Acorda cansado mesmo depois de dormir? Isso pode estar sobrecarregando o seu coração.",
     "Reels ≤60s. Abrir com sintomas que o paciente reconhece ('ronco, acordar sem energia'). Mecanismo simples: falta de oxigênio à noite → PA sobe → coração sofre. Citar polissonografia. CTA: 'Pergunta pro seu parceiro se você ronca e agenda uma avaliação.'",
     "—", "Agosto"),

    (25, "Sono & Estresse",
     "Quantas horas de sono protegem o coração?",
     "Post único", "TOPO",
     "Estabelecer sono como pilar da saúde cardiovascular — simples e direto",
     "7 a 9 horas. Não é luxo — é cardioprevenção.",
     "Post único com tipografia limpa e impactante. Legenda: dado sobre sono <6h e risco de infarto. Tom firme mas acolhedor. CTA: 'Compartilha com quem dorme pouco achando que tá tudo bem.'",
     "—", "Agosto"),

    (26, "Sono & Estresse",
     "Síndrome de Takotsubo: o coração pode se partir de emoção?",
     "Reels", "MEIO",
     "Explicar o Takotsubo de forma emocional e educativa — alta compartilhabilidade",
     "Dra., é verdade que existe síndrome do coração partido? Uma paciente me fez essa pergunta — e a resposta é sim.",
     "Reels ≤60s. Abrir com a pergunta real da paciente. Explicar Takotsubo em linguagem leiga: 'o estresse emocional intenso pode paralisar parte do coração temporariamente'. Quem está em risco (mulheres pós-menopausa). Como tratar. CTA: 'Compartilha com quem está passando por um momento difícil.'",
     "Pergunte à Dra. Paôla 🩺", "Fevereiro"),

    (27, "Sono & Estresse",
     "Estresse crônico deixa marcas reais no coração",
     "Carrossel", "MEIO",
     "Aprofundar a relação entre estresse, inflamação e doença cardiovascular",
     "O estresse não é 'coisa da cabeça'. Ele deixa marcas reais e visíveis no seu coração.",
     "Carrossel 6 slides. S1: gancho debunking. S2: mecanismo (cortisol → inflamação → aterosclerose) com visual simples. S3: estresse agudo vs crônico. S4–5: estratégias com evidência (exercício, sono, respiração). S6: CTA 'Salva esse post — pode ser a conversa mais importante que você tem com alguém hoje.'",
     "—", "Maio"),

    # ── EXAMES & DIAGNÓSTICO (7 temas) ────────────────────────────────────────
    (28, "Exames & Diagnóstico",
     "ECG: o que essa folha de papel revela sobre você",
     "Carrossel", "MEIO",
     "Desmistificar o ECG e mostrar seu valor diagnóstico de forma visual e acessível",
     "Esse exame é só uma folha de papel. Mas ela conta a história do seu coração — e eu vou te ensinar a ler.",
     "Carrossel 7 slides. S1: gancho. S2: foto real de ECG com destaque nas ondas principais. S3–5: o que cada parte revela (ritmo, frequência, isquemia) em linguagem leiga. S6: quando o médico pede. S7: CTA 'Salva — na próxima consulta você vai entender seu próprio ECG.'",
     "—", "Agosto"),

    (29, "Exames & Diagnóstico",
     "Ecocardiograma: por que o médico pediu esse exame?",
     "Reels", "MEIO",
     "Explicar o ecocardiograma de forma acessível para pacientes que receberam a solicitação",
     "O cardiologista pediu um ecocardiograma. O que isso significa? Primeiro: não precisa se assustar.",
     "Reels ≤60s. Abrir tranquilizando. Explicar: 'é um ultrassom do coração, não dói, não irradia'. O que avalia (estrutura, função, válvulas). Quando é indicado. CTA: 'Deixa nos comentários se você já fez um.'",
     "Pergunte à Dra. Paôla 🩺", "Setembro"),

    (30, "Exames & Diagnóstico",
     "MAPA: quando a pressão no consultório engana",
     "Post único", "TOPO",
     "Explicar o MAPA e o fenômeno da 'hipertensão do jaleco branco' de forma direta",
     "Você sabia que a sua pressão pode ser diferente em casa e no consultório? Isso tem nome — e tem solução.",
     "Post único com foto do aparelho + legenda explicativa. Conceito de 'hipertensão do jaleco branco' em 2 frases. Quando o MAPA é indicado. CTA: 'Pergunta pro seu cardiologista se você precisa desse exame.'",
     "—", "Abril"),

    (31, "Exames & Diagnóstico",
     "Teste ergométrico: por que o médico pediu a esteira?",
     "Reels", "MEIO",
     "Orientar sobre o que o teste ergométrico avalia e quando é indicado",
     "O cardiologista pediu uma esteira. O que esse exame realmente mede sobre o seu coração?",
     "Reels ≤60s. Abrir com a dúvida do paciente. Explicar: capacidade funcional, isquemia e arritmia durante esforço. Para quem é indicado. Como é feito. CTA: 'Já fez esse exame? Conta nos comentários como foi.'",
     "Pergunte à Dra. Paôla 🩺", "Novembro"),

    (32, "Exames & Diagnóstico",
     "Checkup cardiovascular: quem precisa e com qual frequência?",
     "Carrossel", "TOPO",
     "Gerar demanda por consulta de rastreamento — mostrar que prevenção é para todos",
     "Quando foi a última vez que você checou a saúde do seu coração? Se você precisou pensar — já passou da hora.",
     "Carrossel 5 slides. S1: gancho com dado de mortalidade cardiovascular. S2: quem deve fazer (grupos de risco + assintomáticos acima de 40). S3: frequência recomendada por faixa etária. S4: o que é avaliado. S5: CTA fundo de funil direto: 'Agende sua avaliação — atendo no Einstein, São Luiz Itaim e Dante Pazzanese.'",
     "—", "Janeiro"),

    (33, "Exames & Diagnóstico",
     "Exame de sangue e coração: o que pedir ao médico",
     "Carrossel", "TOPO",
     "Empoderar o paciente a saber quais exames solicitar para rastreamento cardiovascular",
     "Você vai ao médico mas não sabe o que pedir pra proteger o seu coração. Aqui está o checklist.",
     "Carrossel tipo checklist. S1: gancho empoderador. S2–5: exames essenciais (colesterol fracionado, glicemia, PCR-us, hemograma) com o que cada um revela em 1 linha. S6: CTA 'Salva esse post antes da sua próxima consulta.'",
     "—", "Abril"),

    (34, "Exames & Diagnóstico",
     "Holter 24h: quando o coração precisa ser monitorado",
     "Post único", "MEIO",
     "Explicar o Holter de forma simples para pacientes que receberão a solicitação",
     "Sentiu o coração disparar ou pular um batimento? Esse exame pode finalmente dar uma resposta.",
     "Post único com foto discreta do aparelho. Legenda: o que o Holter registra, quando é indicado (palpitações, tontura, síncope). Tom: tranquilizador, não alarmista. CTA: 'Já usou um Holter? Conta nos comentários.'",
     "Pergunte à Dra. Paôla 🩺", "Julho"),

    # ── DOENÇAS CARDIOVASCULARES (6 temas) ────────────────────────────────────
    (35, "Doenças Cardiovasculares",
     "Insuficiência cardíaca: os sintomas que você não pode ignorar",
     "Reels", "MEIO",
     "Alertar para sintomas de IC e motivar busca por diagnóstico precoce",
     "Cansaço ao subir uma escada. Falta de ar ao deitar. Pés inchados. Esses sinais têm algo em comum — e falam sobre o seu coração.",
     "Reels ≤60s. Abrir com sintomas concretos que o paciente reconhece. Sequência emocional: 'muita gente acha que é cansaço de trabalho ou da idade'. Revelar: podem ser sinais de IC. CTA: 'Encaminha pra quem tem esses sintomas e ainda não foi ao médico.'",
     "—", "Outubro"),

    (36, "Doenças Cardiovasculares",
     "Arritmia: quando o coração bate diferente",
     "Reels", "MEIO",
     "Tranquilizar e orientar quem já sentiu arritmia sobre quando buscar avaliação urgente",
     "Senti o coração bater diferente. Preciso me preocupar? Depende — e eu vou te explicar a diferença.",
     "Reels ≤60s. Abrir com a dúvida/medo do paciente. Distinguir em linguagem leiga: benignas (extrassístoles comuns) vs sinais de alerta (taquicardia sustentada, síncope). Quando ir ao pronto-socorro vs consulta eletiva. CTA: 'Salva esse post — pode ser útil num momento de susto.'",
     "Pergunte à Dra. Paôla 🩺", "Maio"),

    (37, "Doenças Cardiovasculares",
     "Miocardiopatia: quando o músculo do coração adoece",
     "Reels", "MEIO",
     "Explicar o conceito de miocardiopatia e seus tipos para pacientes com diagnóstico ou risco",
     "Você sabia que o próprio músculo do coração pode adoecer — independente dos vasos e das válvulas?",
     "Reels ≤60s. Abrir com dado surpresa. Tipos em linguagem leiga: 'dilatada (coração grande e fraco), hipertrófica (músculo espessado), restritiva (rígido)'. Conectar com IC avançada. CTA: 'Agende avaliação se tem histórico familiar ou sintomas.'",
     "—", "Outubro"),

    (38, "Doenças Cardiovasculares",
     "Aterosclerose: a placa que se forma em silêncio",
     "Carrossel", "TOPO",
     "Visualizar o processo de aterosclerose para aumentar percepção de risco e motivar prevenção",
     "A placa de gordura no seu vaso não surge do dia para a noite. É um processo de anos — que começa antes dos 30.",
     "Carrossel visual com ilustração do processo. S1: gancho temporal. S2: artéria saudável. S3: formação da placa (linguagem: 'o colesterol se instala nas paredes'). S4: estreitamento e risco de infarto. S5: o que freia esse processo. S6: CTA 'Salva — e compartilha com quem acha que ainda é jovem demais pra se preocupar.'",
     "—", "Julho"),

    (39, "Doenças Cardiovasculares",
     "Infarto: os sinais que muita gente não reconhece",
     "Reels", "TOPO",
     "Ampliar capacidade do público de reconhecer sintomas atípicos de infarto e agir rápido",
     "Infarto não começa sempre com dor no peito. Esses outros sinais salvam vidas — e muita gente não conhece.",
     "Reels de alto impacto. Listar sintomas atípicos: dor no braço/mandíbula/costas, náusea, suor frio. Atenção especial: sintomas em mulheres são diferentes. Instrução direta: 'Ligue 192 imediatamente'. Alta compartilhabilidade. CTA: 'Compartilha. Pode salvar uma vida.'",
     "—", "Setembro"),

    (40, "Doenças Cardiovasculares",
     "Prevenção primária vs secundária: você sabe qual é a sua?",
     "Post único", "TOPO",
     "Explicar os dois tipos de prevenção e posicionar a Paôla como especialista em ambos",
     "Mito ou Verdade: prevenção cardiovascular é só para quem ainda não teve problema no coração.",
     "Post único formato Mito ou Verdade. Resposta: MITO. Legenda: prevenção primária (antes da doença) e secundária (após evento) têm estratégias diferentes — ambas são tratamento. CTA fundo de funil: 'Sabe qual é a sua? Agende sua avaliação.'",
     "Mito ou Verdade? 🫀", "Abril"),
]

# ── COLUNAS ────────────────────────────────────────────────────────────────────
headers = [
    "#",
    "LINHA EDITORIAL",
    "SUBTEMA",
    "TEMA / TÍTULO DO POST",
    "FORMATO",
    "ETAPA DO FUNIL",
    "OBJETIVO DA POSTAGEM",
    "GANCHO DE ABERTURA",
    "ORIENTAÇÕES DE PRODUÇÃO",
    "SÉRIE RECORRENTE",
    "MÊS SUGERIDO",
]
col_widths = [5, 24, 22, 46, 14, 16, 50, 58, 68, 26, 16]

# ── HELPERS ───────────────────────────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
border_light = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color, bold=False, size=10):
    return Font(name="Calibri", color=hex_color, bold=bold, size=size)

def align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="top", wrap_text=wrap)

# ── LINHA 1: TÍTULO ───────────────────────────────────────────────────────────
ws.merge_cells("A1:K1")
c = ws["A1"]
c.value = "CALENDÁRIO EDITORIAL 2026 — MEDICINA DO ESTILO DE VIDA"
c.fill = fill(HEADER_BG)
c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 38

# ── LINHA 2: SUBTÍTULO ────────────────────────────────────────────────────────
ws.merge_cells("A2:K2")
c = ws["A2"]
c.value = "@paolapreto · Instagram · Linha Editorial 1 de 4 · 40 temas · Tom de voz: amiga médica, acolhedora e direta"
c.fill = fill("2E6B8A")
c.font = Font(name="Calibri", color="FFFFFF", size=10, italic=True)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# ── LINHA 3: CABEÇALHOS ───────────────────────────────────────────────────────
ws.row_dimensions[3].height = 34
for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
    c = ws.cell(row=3, column=ci, value=h)
    c.fill = fill("2E6B8A")
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(
        left=Side(style="medium", color="FFFFFF"),
        right=Side(style="medium", color="FFFFFF"),
        bottom=Side(style="medium", color="FFFFFF"),
    )
    ws.column_dimensions[get_column_letter(ci)].width = w

# ── DADOS ─────────────────────────────────────────────────────────────────────
LINHA_EDITORIAL = "Medicina do Estilo de Vida"

prev_subtema = None
for ri, (num, subtema, tema, fmt, funil, objetivo, gancho, orient, serie, mes) in enumerate(topics, start=4):
    ws.row_dimensions[ri].height = 62

    sub_bg, sub_fg = SUBTEMA_CORES.get(subtema, ("F0F0F0", "333333"))
    fun_bg, fun_fg = FUNIL_CORES.get(funil, ("FFFFFF", "333333"))
    fmt_bg, fmt_fg = FORMATO_CORES.get(fmt, ("F0F0F0", "333333"))
    ser_bg, ser_fg = SERIE_CORES.get(serie, ("F8F9FA", "6C757D"))

    row_data = [
        (str(num),         fill("F8F9FA"),   font("555555", bold=True),  "center"),
        (LINHA_EDITORIAL,  fill("EAF4FB"),   font("1A5276"),             "center"),
        (subtema,          fill(sub_bg),     font(sub_fg, bold=True),    "center"),
        (tema,             fill("FFFFFF"),   font("1A1A2E", bold=True),  "left"),
        (fmt,              fill(fmt_bg),     font(fmt_fg, bold=True),    "center"),
        (funil,            fill(fun_bg),     font(fun_fg, bold=True),    "center"),
        (objetivo,         fill("FDFEFE"),   font("2C3E50"),             "left"),
        (gancho,           fill("FFFDE7"),   font("5D4037"),             "left"),
        (orient,           fill("F9F9F9"),   font("4A4A4A"),             "left"),
        (serie,            fill(ser_bg),     font(ser_fg, bold=True),    "center"),
        (mes,              fill("EAF4FB"),   font("1A5276", bold=True),  "center"),
    ]

    for ci, (val, f, fo, al) in enumerate(row_data, start=1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill = f
        c.font = fo
        c.alignment = align(al)
        c.border = border_light

    # borda superior mais grossa ao mudar de subtema
    if subtema != prev_subtema:
        prev_subtema = subtema
        for ci in range(1, 12):
            c = ws.cell(row=ri, column=ci)
            c.border = Border(
                left=c.border.left, right=c.border.right,
                bottom=c.border.bottom,
                top=Side(style="medium", color="888888"),
            )

# ── LEGENDA ───────────────────────────────────────────────────────────────────
lr = len(topics) + 5

def legend_section(title, items, start_row):
    ws.merge_cells(f"A{start_row}:K{start_row}")
    c = ws[f"A{start_row}"]
    c.value = title
    c.fill = fill(HEADER_BG)
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 22
    r = start_row + 1
    ws.row_dimensions[r].height = 22
    for i, (label, (bg, fg)) in enumerate(items.items()):
        col = i * 2 + 1
        cell = ws.cell(row=r, column=col, value=label)
        cell.fill = fill(bg)
        cell.font = Font(name="Calibri", color=fg, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return r + 2

lr = legend_section("SUBTEMA",         SUBTEMA_CORES,  lr)
lr = legend_section("ETAPA DO FUNIL",  FUNIL_CORES,    lr)
lr = legend_section("FORMATO",         FORMATO_CORES,  lr)
lr = legend_section("SÉRIE RECORRENTE",SERIE_CORES,    lr)

# ── CONGELAR CABEÇALHO + FILTRO ───────────────────────────────────────────────
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:K{len(topics) + 3}"

# ── SALVAR ─────────────────────────────────────────────────────────────────────
out = "/Users/lucas/Downloads/Paôla - Calendário Editorial 2026/Calendário Editorial 2026 — Medicina do Estilo de Vida.xlsx"
wb.save(out)
print(f"Salvo em: {out}")
