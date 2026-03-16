"""
Gera a aba "Conversão" e regenera "Objetivos Mês a Mês" com:
  - Posts planejados atualizados (com os temas de Conversão definidos)
  - 2 novas colunas: Estratégia de Anúncio + KPIs Prioritários
  - Seção de guia de campanhas ao final da aba de objetivos
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = (
    "/Users/lucas/Downloads/Paôla - Calendário Editorial 2026/"
    "Calendário Editorial 2026 — Medicina do Estilo de Vida.xlsx"
)

wb = load_workbook(XLSX)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS COMPARTILHADOS
# ══════════════════════════════════════════════════════════════════════════════
thin  = Side(style="thin",   color="CCCCCC")
thick = Side(style="medium", color="888888")
border_light = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fnt(hex_color, bold=False, size=10, italic=False):
    return Font(name="Calibri", color=hex_color, bold=bold, size=size, italic=italic)

def aln(h="left", wrap=True, v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_title(ws, span, row, value, bg, fg, size=14, bold=True, height=38):
    ws.merge_cells(f"A{row}:{span}{row}")
    c = ws[f"A{row}"]
    c.value = value
    c.fill = fill(bg)
    c.font = Font(name="Calibri", color=fg, bold=bold, size=size)
    c.alignment = aln("center", v="center")
    ws.row_dimensions[row].height = height

def top_border(ws, row, ncols):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row, column=ci)
        c.border = Border(
            left=c.border.left, right=c.border.right, bottom=c.border.bottom,
            top=thick,
        )


# ══════════════════════════════════════════════════════════════════════════════
# 1. ABA: CONVERSÃO
# ══════════════════════════════════════════════════════════════════════════════

if "Conversão" in wb.sheetnames:
    del wb["Conversão"]

# Insere como segunda aba (depois de Objetivos)
idx = wb.sheetnames.index("Objetivos Mês a Mês") + 1 if "Objetivos Mês a Mês" in wb.sheetnames else 1
wc = wb.create_sheet("Conversão", idx)

# ── Paleta ────────────────────────────────────────────────────────────────────
C_HEADER    = "3D1212"
C_SUBHEADER = "7B1A1A"
C_LINHA     = "FDF2F2"

SUBTEMA_CORES_C = {
    "Apresentação & Posicionamento": ("FDEDEC", "922B21"),
    "Diferenciais & Especialidade":  ("F4ECF7", "76448A"),
    "Onde Atendo & Agenda":          ("FDEBD0", "784212"),
    "Primeira Consulta":             ("D5F5E3", "1E8449"),
    "Q&A & Engajamento":             ("FEF9E7", "9A7D0A"),
}

FUNIL_CORES_C = {
    "FUNDO": ("FDEDEC", "922B21"),
}

FORMATO_CORES_C = {
    "Reels":      ("E8DAEF", "6C3483"),
    "Carrossel":  ("D6EAF8", "1A5276"),
    "Post único": ("D5F5E3", "1E8449"),
}

SERIE_CORES_C = {
    "Pergunte à Dra. Paôla 🩺": ("D1ECF1", "0C5460"),
    "—":                         ("F8F9FA", "6C757D"),
}

# ── Temas de Conversão ────────────────────────────────────────────────────────
# #, SUBTEMA, TEMA/TÍTULO, FORMATO, FUNIL, OBJETIVO, GANCHO, ORIENTAÇÕES, SÉRIE, MÊS
conv_topics = [
    (
        1,
        "Apresentação & Posicionamento",
        "Apresentação: nova fase em São Paulo",
        "Reels", "FUNDO",
        "Marcar o início do perfil renovado apresentando a Paôla como especialista em IC avançada e transplante, "
        "recém estabelecida em SP, e humanizando a marca sem perder autoridade médica.",
        "Eu sou cardiologista. Mas a especialidade que escolhi cuida de quem o coração já não consegue mais sozinho. "
        "Agora em São Paulo — e quero te apresentar o meu trabalho.",
        "Reels ≤60s. Abertura com afirmação de posicionamento (não 'olá, me chamo'). Mostrar rosto, tom acolhedor e direto. "
        "Citar os 3 locais de atendimento. Encerrar com CTA claro: 'Me segue — toda semana tem conteúdo sobre saúde do coração.' "
        "Produção simples: fundo neutro ou consultório. Legenda reforça apresentação + link na bio.",
        "—",
        "Abril",
    ),
    (
        2,
        "Diferenciais & Especialidade",
        "Meus diferenciais: IC avançada e transplante cardíaco",
        "Carrossel", "FUNDO",
        "Explicar de forma acessível o que é a especialidade de IC avançada e transplante, e por que ter "
        "um especialista faz diferença — gerando desejo de consulta em quem já tem diagnóstico ou suspeita.",
        "Nem todo cardiologista trata o mesmo. Vou te explicar o que faz um especialista em insuficiência cardíaca avançada "
        "— e por que isso pode mudar tudo no seu cuidado.",
        "Carrossel 6 slides. S1: gancho de diferenciação. S2: o que é IC avançada em 1 frase leiga "
        "('quando o coração não consegue mais bombear o suficiente'). S3: o que o especialista faz diferente "
        "(gestão do Quarteto Fantástico, avaliação para transplante). S4: perfil do paciente ideal. "
        "S5: onde atendo (Einstein, São Luiz Itaim, Dante Pazzanese). S6: CTA direto 'Agende sua avaliação — link na bio.'",
        "—",
        "Abril",
    ),
    (
        3,
        "Apresentação & Posicionamento",
        "Cuidar da saúde = o melhor presente para quem você ama",
        "Post único", "FUNDO",
        "Usar o gancho emocional do Dia das Mães para conectar cuidado com saúde e gerar CTA de agendamento "
        "sem parecer apelativo — transformar presença em valor.",
        "A melhor coisa que você pode fazer pela sua família não tem embrulho. É continuar aqui.",
        "Post único com visual emocional (cores suaves, tipografia impactante). Legenda curta: 2-3 frases "
        "conectando saúde cardiovascular à presença para a família. Encerrar com CTA: 'Agende sua consulta — "
        "atendo no Einstein, São Luiz Itaim e Dante Pazzanese.' Publicar próximo ao Dia das Mães (10/mai).",
        "—",
        "Maio",
    ),
    (
        4,
        "Onde Atendo & Agenda",
        "Onde te atendo em SP e como agendar sua consulta",
        "Post único", "FUNDO",
        "Remover a fricção do processo de agendamento: dar ao seguidor todas as informações práticas "
        "de forma clara e memorável, eliminando a barreira 'não sei como marcar'.",
        "Você quer agendar, mas não sabe por onde começar. Deixa eu facilitar isso pra você.",
        "Post único estilo guia prático. Visual com os 3 locais (ícones ou fotos). Legenda: nome dos hospitais, "
        "como agendar em cada um (telefone, site, aplicativo). Tom: acolhedor e prático. "
        "CTA: 'Salva esse post — quando precisar, estará aqui.' Fixar esse post no perfil após publicação.",
        "—",
        "Junho",
    ),
    (
        5,
        "Primeira Consulta",
        "O que esperar da sua primeira consulta comigo",
        "Carrossel", "FUNDO",
        "Reduzir a ansiedade do paciente sobre o que acontece na consulta e antecipar o valor da avaliação "
        "especializada — transformando curiosidade em decisão de agendamento.",
        "Dra., o que você faz na consulta? Essa pergunta chega toda semana. Vou responder de um jeito diferente: "
        "te conto exatamente o que acontece quando você senta na minha frente.",
        "Carrossel 6 slides. S1: gancho com a pergunta do paciente. S2: 'começo pela sua história' — "
        "anamnese humanizada. S3: exame físico cardiovascular (o que avalio). S4: revisão dos exames "
        "(o que trazer). S5: construção do plano junto com o paciente. S6: CTA 'Pronto para agendar? "
        "Link na bio — atendo no Einstein, São Luiz Itaim e Dante Pazzanese.'",
        "Pergunte à Dra. Paôla 🩺",
        "Julho",
    ),
    (
        6,
        "Diferenciais & Especialidade",
        "Por que consultar um especialista em IC avançada?",
        "Reels", "FUNDO",
        "Gerar consciência sobre a importância do acompanhamento especializado em IC — e posicionar a Paôla "
        "como a profissional certa para esse público. Alcançar pacientes com IC já diagnosticada.",
        "Se você tem insuficiência cardíaca e não está sendo acompanhado por um especialista, precisa ouvir isso.",
        "Reels ≤60s. Tom direto e urgente (sem alarmismo). Dado de impacto: mortalidade de IC não tratada "
        "adequadamente. Diferenciar clínico geral vs cardiologista vs especialista em IC. "
        "O que o especialista monitora (FEVE, volume, ajuste do Quarteto). CTA: 'Agende sua avaliação — "
        "link na bio.'",
        "—",
        "Agosto",
    ),
    (
        7,
        "Onde Atendo & Agenda",
        "Dia Mundial do Coração: agende sua avaliação cardiovascular",
        "Post único", "FUNDO",
        "Aproveitar a data simbólica do Dia Mundial do Coração (29/set) para gerar CTA de agendamento "
        "com urgência legítima e emocional — pico de intenção do público nessa data.",
        "Hoje é o Dia Mundial do Coração. E eu tenho um pedido: não deixa pra amanhã o que pode salvar a sua vida.",
        "Post único de alto impacto visual (coração, paleta vermelha). Legenda curta e emocional: "
        "estatística de mortalidade cardiovascular + convite à consulta. CTA direto: "
        "'Agende agora — Einstein, São Luiz Itaim, Dante Pazzanese. Link na bio.' "
        "Publicar em 29/set. Potencial de alto alcance orgânico — considerar impulsionamento.",
        "—",
        "Setembro",
    ),
    (
        8,
        "Q&A & Engajamento",
        "Q&A: respondendo suas dúvidas sobre o coração",
        "Carrossel", "FUNDO",
        "Gerar engajamento alto com perguntas reais dos seguidores, mostrar acessibilidade e construir "
        "confiança — transformando seguidores engajados em pacientes.",
        "Toda semana chegam perguntas na minha caixa. Hoje eu respondo as mais repetidas — com a mesma "
        "atenção que daria no consultório.",
        "Carrossel 6-8 slides. Cada slide = 1 pergunta real + resposta curta em linguagem leiga. "
        "Formato: fundo escuro, pergunta em branco, resposta em destaque. Selecionar perguntas de alto "
        "interesse geral (ex: 'posso tomar anticoncepcional com pressão alta?', 'palpitação é perigosa?'). "
        "Último slide: 'Manda sua pergunta nos comentários — responderei no próximo Q&A.' "
        "CTA de agendamento discreto no final.",
        "Pergunte à Dra. Paôla 🩺",
        "Outubro",
    ),
    (
        9,
        "Diferenciais & Especialidade",
        "Por que me escolher como sua cardiologista",
        "Reels", "FUNDO",
        "Post de autoridade e diferenciação: consolidar a percepção de valor da Paôla antes do "
        "fechamento do ano, capturando pacientes indecisos que já a seguem há meses.",
        "Existem muitos cardiologistas em São Paulo. Vou te contar o que me diferencia — não para me vender, "
        "mas para que você tome a melhor decisão para a sua saúde.",
        "Reels ≤60s. Tom seguro, sem arrogância. Citar: formação em IC avançada e transplante, "
        "hospitais de referência (Einstein, São Luiz, Dante Pazzanese), abordagem de medicina do estilo "
        "de vida + alta tecnologia, atendimento humanizado. Não citar preços (vedado pelo CFM). "
        "CTA: 'Agende sua consulta — link na bio.'",
        "—",
        "Novembro",
    ),
    (
        10,
        "Onde Atendo & Agenda",
        "Agende seu checkup cardiovascular para 2027",
        "Post único", "FUNDO",
        "Fechar o ano com um CTA de planejamento preventivo, criando senso de urgência positivo "
        "('ano novo, coração avaliado') e gerando agenda para janeiro/2027.",
        "O ano acabou. Mas o coração trabalhou 365 dias sem parar. Quando foi a última vez que você cuidou dele?",
        "Post único com visual de virada de ano (minimalista, não kitsch). Legenda: 1 dado sobre "
        "prevenção cardiovascular + convite ao checkup. CTA: 'Agende agora para janeiro — "
        "atendo no Einstein, São Luiz Itaim e Dante Pazzanese. Link na bio.' "
        "Considerar impulsionamento para alcançar quem ainda não é seguidor.",
        "—",
        "Dezembro",
    ),

    # ── TEMAS BANCO (sem mês fixo — reserva e substituição) ───────────────────
    (
        11,
        "Apresentação & Posicionamento",
        "Minha trajetória: por que escolhi IC avançada e transplante",
        "Reels", "FUNDO",
        "Humanizar a marca médica contando a história de como chegou à subespecialidade mais complexa da "
        "cardiologia — gerar conexão e confiança com pacientes que precisam de referência em IC.",
        "A maioria dos médicos foge da IC avançada. Eu corri em direção a ela. Deixa eu te contar por quê.",
        "Reels ≤90s. Tom pessoal e autêntico — não um currículo, uma história. Contar o que levou à escolha: "
        "o primeiro contato com IC avançada, o que a especialidade representa. Encerrar com propósito: "
        "'Escolhi cuidar de quem o coração não consegue mais sozinho.' "
        "CTA: 'Me segue — estou aqui pra te mostrar que IC avançada tem tratamento.'",
        "—",
        "Banco",
    ),
    (
        12,
        "Diferenciais & Especialidade",
        "IC avançada não é o fim — é o começo de um novo protocolo",
        "Reels", "FUNDO",
        "Desconstruir o medo e o estigma associados ao diagnóstico de IC avançada, posicionando o "
        "tratamento especializado como esperança real e mudança de trajetória.",
        "Meu paciente chegou achando que tinha recebido uma sentença. O que eu disse a ele mudou tudo.",
        "Reels ≤60s. Abrir com narrativa de paciente (sem identificação). Virada emocional: o que IC avançada "
        "significa hoje com o Quarteto Fantástico + possibilidade de transplante. "
        "Tom: esperançoso e realista — sem prometer cura. "
        "CTA: 'Se você ou alguém próximo tem IC avançada, me manda uma mensagem.'",
        "—",
        "Banco",
    ),
    (
        13,
        "Primeira Consulta",
        "Quais exames trazer na primeira consulta comigo?",
        "Carrossel", "FUNDO",
        "Reduzir a barreira de agendamento ao dar um checklist prático do que trazer — elimina a "
        "insegurança de 'não sei se estou preparado para a consulta'.",
        "Antes de vir ao consultório, você precisa ler isso. Vai tornar nossa consulta muito mais completa.",
        "Carrossel 6 slides estilo checklist. S1: gancho + 'guarda esse post'. "
        "S2: exames cardiovasculares (ECG, eco, MAPA, Holter se tiver). "
        "S3: exames de sangue (colesterol, glicemia, função renal, TSH). "
        "S4: laudos de imagem e relatórios anteriores. "
        "S5: lista de medicamentos em uso (nome + dose). "
        "S6: CTA 'Pronto para agendar? Link na bio — Einstein, São Luiz Itaim, Dante Pazzanese.' "
        "Indicado para fixar no perfil.",
        "—",
        "Banco",
    ),
    (
        14,
        "Primeira Consulta",
        "Como me preparar para a consulta cardiológica",
        "Post único", "FUNDO",
        "Empoderar o paciente novo, reduzindo ansiedade pré-consulta e aumentando a qualidade do "
        "encontro clínico desde o primeiro contato.",
        "Uma consulta boa começa antes de você entrar no consultório. Aqui está o que fazer.",
        "Post único estilo guia visual rápido. 4-5 dicas práticas: anotar sintomas com frequência e "
        "intensidade, trazer lista completa de medicamentos, registrar a PA em casa se hipertenso, "
        "não ir em jejum (a não ser que haja exame), anotar perguntas com antecedência. "
        "Tom tranquilizador — remove o medo de 'não saber o que fazer'. "
        "CTA: 'Salva esse post antes da sua próxima consulta.'",
        "—",
        "Banco",
    ),
    (
        15,
        "Onde Atendo & Agenda",
        "Atendo convênio? Como acessar minha consulta",
        "Post único", "FUNDO",
        "Remover a barreira de acesso mais frequente: 'será que aceita meu plano?' — ser transparente "
        "sobre as opções de agendamento e eliminar a desistência por dúvida prática.",
        "A pergunta que mais recebo no DM: Dra., você atende convênio? Vou responder de uma vez por todas.",
        "Post único direto e prático. Listar os planos aceitos em cada hospital "
        "(preencher com dados reais antes de publicar). Informar canal de agendamento de cada hospital "
        "(telefone, site, aplicativo). Mencionar opção de consulta particular quando aplicável. "
        "Tom: transparente e acessível. "
        "CTA: 'Me manda uma mensagem se tiver dúvida sobre seu plano específico.'",
        "—",
        "Banco",
    ),
    (
        16,
        "Onde Atendo & Agenda",
        "Einstein, São Luiz Itaim, Dante Pazzanese: por que esses hospitais?",
        "Carrossel", "FUNDO",
        "Posicionar os hospitais como referências de excelência em cardiologia e reforçar que a escolha "
        "não é arbitrária — aumentar a confiança e o senso de que o paciente está em boas mãos.",
        "Não escolhi esses hospitais por acaso. Cada um tem um papel no meu trabalho com você.",
        "Carrossel 4 slides. S1: gancho com afirmação de posicionamento. "
        "S2: Hospital Albert Einstein — referência em IC avançada e transplante, centro de excelência. "
        "S3: São Luiz Itaim — excelência cardiovascular, equipe multidisciplinar, localização estratégica. "
        "S4: Dante Pazzanese — instituto cardiológico de referência nacional, histórico e trajetória da Paôla. "
        "Último slide: mapa ou logos dos hospitais + CTA de agendamento com link na bio.",
        "—",
        "Banco",
    ),
    (
        17,
        "Diferenciais & Especialidade",
        "Acompanhamento pós-transplante: o que acontece depois?",
        "Reels", "FUNDO",
        "Mostrar o ciclo completo do cuidado pós-transplante — desmistificar e posicionar a Paôla como "
        "especialista em todo o espectro do tratamento, do diagnóstico ao pós-cirúrgico.",
        "O transplante foi. E agora? O que ninguém te conta sobre o que vem depois.",
        "Reels ≤60s. Abrir com a lacuna de informação que pacientes e familiares sentem. "
        "Explicar em linguagem leiga: biópsia de rotina (o que é e por que importa), ajuste de "
        "imunossupressores, restrições iniciais e liberdades progressivas, qualidade de vida real a longo prazo. "
        "Tom: realista, esperançoso e prático. "
        "CTA: 'Manda nos comentários se você é transplantado ou familiar de transplantado.'",
        "—",
        "Banco",
    ),
    (
        18,
        "Q&A & Engajamento",
        "Quando buscar um segundo parecer cardiológico?",
        "Reels", "FUNDO",
        "Empoderar o paciente a buscar especialização quando necessário — posicionando a Paôla como "
        "a referência ideal para segundo parecer em IC avançada.",
        "Você tem o direito de buscar um segundo parecer. E às vezes ele muda tudo.",
        "Reels ≤60s. Abrir com dado ou caso (sem identificação) em que o segundo parecer foi decisivo. "
        "Quando buscar: IC sem melhora com tratamento, sintomas inexplicados, cirurgia cardíaca indicada, "
        "dúvida sobre diagnóstico. Tom empoderador — sem criticar outros médicos. "
        "CTA: 'Se você está em dúvida, me manda uma mensagem — posso ser o seu segundo parecer.'",
        "Pergunte à Dra. Paôla 🩺",
        "Banco",
    ),
    (
        19,
        "Q&A & Engajamento",
        "Perguntas que você tem direito de fazer ao seu médico",
        "Carrossel", "FUNDO",
        "Posicionar a Paôla como médica que respeita a autonomia do paciente e estimula participação ativa "
        "— diferencial de relacionamento que converte seguidores em pacientes que se sentem seguros.",
        "Consulta não é monólogo. Você tem perguntas importantes a fazer — e o direito de receber respostas.",
        "Carrossel 7 slides. S1: gancho. "
        "S2-6: perguntas que todo paciente deveria fazer: 'Por que esse medicamento?' / "
        "'Quais efeitos colaterais devo monitorar?' / 'Quais sinais de alerta me fazem ligar agora?' / "
        "'Quando retorno?' / 'Existe outra opção de tratamento?'. "
        "S7: 'Na minha consulta, você pode e deve fazer todas essas perguntas.' "
        "CTA: 'Salva e leva para sua próxima consulta.'",
        "—",
        "Banco",
    ),
    (
        20,
        "Diferenciais & Especialidade",
        "Clínico geral, cardiologista, especialista em IC: qual é a diferença?",
        "Carrossel", "FUNDO",
        "Educar sobre os diferentes papéis no cuidado cardiovascular, criando consciência de que o "
        "paciente com IC avançada precisa de acompanhamento especializado — e gerando demanda de consulta.",
        "Clínico geral, cardiologista, especialista em IC avançada. São três coisas diferentes — "
        "e o seu coração merece saber a diferença.",
        "Carrossel comparativo 5 slides. S1: gancho de diferenciação. "
        "S2: clínico geral — rastreamento, controle de fatores de risco, encaminhamento. "
        "S3: cardiologista geral — diagnóstico e tratamento da maioria das condições cardíacas. "
        "S4: especialista em IC avançada — gestão do Quarteto Fantástico, avaliação para dispositivos "
        "(CDI, ressincronizador), indicação de transplante, protocolo intensivo. "
        "S5: quando cada um é necessário + CTA: "
        "'Se você tem IC, me segue — estou aqui para esse acompanhamento especializado.'",
        "—",
        "Banco",
    ),
]

# ── Estrutura da planilha ─────────────────────────────────────────────────────
headers_c = [
    "#", "SUBTEMA", "TEMA / TÍTULO DO POST", "FORMATO",
    "ETAPA DO FUNIL", "OBJETIVO DA POSTAGEM",
    "GANCHO DE ABERTURA", "ORIENTAÇÕES DE PRODUÇÃO", "SÉRIE RECORRENTE", "MÊS SUGERIDO",
]
col_widths_c = [5, 26, 44, 14, 16, 52, 58, 68, 26, 16]

for i, w in enumerate(col_widths_c, 1):
    wc.column_dimensions[get_column_letter(i)].width = w

# Título
set_title(wc, "J", 1,
    "CALENDÁRIO EDITORIAL 2026 — CONVERSÃO",
    C_HEADER, "FFFFFF", size=14)

# Subtítulo
set_title(wc, "J", 2,
    "@paolapreto · Instagram · Linha Editorial 2 de 4 · 20 temas (10 com mês fixo + 10 banco de reserva) · "
    "Meta: converter seguidor em paciente · Funil: FUNDO",
    C_SUBHEADER, "FFFFFF", size=10, bold=False, height=20)

# Cabeçalhos
wc.row_dimensions[3].height = 34
for ci, (h, w) in enumerate(zip(headers_c, col_widths_c), 1):
    c = wc.cell(row=3, column=ci, value=h)
    c.fill = fill(C_SUBHEADER)
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    c.alignment = aln("center", wrap=True, v="center")
    c.border = Border(
        left=Side(style="medium", color="FFFFFF"),
        right=Side(style="medium", color="FFFFFF"),
        bottom=Side(style="medium", color="FFFFFF"),
    )

# Dados
LINHA_EDITORIAL_C = "Conversão"
prev_sub = None

for ri, (num, subtema, tema, fmt, funil, objetivo, gancho, orient, serie, mes) in enumerate(conv_topics, start=4):
    wc.row_dimensions[ri].height = 70

    sub_bg, sub_fg = SUBTEMA_CORES_C.get(subtema, ("F0F0F0", "333333"))
    fun_bg, fun_fg = FUNIL_CORES_C.get(funil, ("FDEDEC", "922B21"))
    fmt_bg, fmt_fg = FORMATO_CORES_C.get(fmt, ("F0F0F0", "333333"))
    ser_bg, ser_fg = SERIE_CORES_C.get(serie, ("F8F9FA", "6C757D"))

    row_data = [
        (str(num),        fill("FDF2F2"),   fnt("922B21", bold=True),  "center"),
        (subtema,         fill(sub_bg),     fnt(sub_fg,   bold=True),  "center"),
        (tema,            fill("FFFFFF"),   fnt("1A1A2E", bold=True),  "left"),
        (fmt,             fill(fmt_bg),     fnt(fmt_fg,   bold=True),  "center"),
        (funil,           fill(fun_bg),     fnt(fun_fg,   bold=True),  "center"),
        (objetivo,        fill("FDFEFE"),   fnt("2C3E50"),             "left"),
        (gancho,          fill("FFFDE7"),   fnt("5D4037"),             "left"),
        (orient,          fill("F9F9F9"),   fnt("4A4A4A"),             "left"),
        (serie,           fill(ser_bg),     fnt(ser_fg,   bold=True),  "center"),
        (mes,             fill("FDEBD0"),   fnt("7B1A1A", bold=True),  "center"),
    ]

    for ci, (val, f, fo, al) in enumerate(row_data, 1):
        c = wc.cell(row=ri, column=ci, value=val)
        c.fill = f; c.font = fo
        c.alignment = aln(al)
        c.border = border_light

    if subtema != prev_sub:
        prev_sub = subtema
        top_border(wc, ri, 10)

# Legenda
lr = len(conv_topics) + 5

def legend_sec(ws, title, items, start_row, ncols=10):
    ws.merge_cells(f"A{start_row}:{get_column_letter(ncols)}{start_row}")
    c = ws[f"A{start_row}"]
    c.value = title
    c.fill = fill(C_HEADER)
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    c.alignment = aln("center", v="center")
    ws.row_dimensions[start_row].height = 22
    r = start_row + 1
    ws.row_dimensions[r].height = 22
    for i, (label, (bg, fg)) in enumerate(items.items()):
        col = i * 2 + 1
        cell = ws.cell(row=r, column=col, value=label)
        cell.fill = fill(bg); cell.font = Font(name="Calibri", color=fg, bold=True, size=10)
        cell.alignment = aln("center", v="center")
    return r + 2

lr = legend_sec(wc, "SUBTEMA",         SUBTEMA_CORES_C, lr)
lr = legend_sec(wc, "FORMATO",         FORMATO_CORES_C, lr)
lr = legend_sec(wc, "SÉRIE RECORRENTE",SERIE_CORES_C,   lr)

wc.freeze_panes = "A4"
wc.auto_filter.ref = f"A3:J{len(conv_topics) + 3}"


# ══════════════════════════════════════════════════════════════════════════════
# 2. ABA: OBJETIVOS MÊS A MÊS  (regenerada com 9 colunas + seção de anúncios)
# ══════════════════════════════════════════════════════════════════════════════

if "Objetivos Mês a Mês" in wb.sheetnames:
    del wb["Objetivos Mês a Mês"]

wo = wb.create_sheet("Objetivos Mês a Mês", 0)

HEADER_BG    = "1A3A4A"
SUBHEADER_BG = "2E6B8A"
SEPT_BG      = "FFF3CD"
ADS_BG       = "1B5E20"   # verde escuro para seção de anúncios
ADS_SUB      = "2E7D32"
ROW_ALT      = "F4F9FC"
ROW_BASE     = "FFFFFF"

# Colunas A-I
col_widths_o = [18, 11, 55, 50, 22, 28, 16, 44, 36]
headers_o = [
    "MÊS", "SEQUÊNCIA", "OBJETIVO DO MÊS", "POSTS PLANEJADOS",
    "LINHAS EDITORIAIS", "DATAS IMPORTANTES", "FUNIL DOMINANTE",
    "ESTRATÉGIA DE ANÚNCIO", "KPIs PRIORITÁRIOS",
]
NCOLS = 9

for i, w in enumerate(col_widths_o, 1):
    wo.column_dimensions[get_column_letter(i)].width = w

# ── Dados dos meses ───────────────────────────────────────────────────────────
meses = [
    {
        "mes":      "Abril",
        "num":      "1º mês",
        "objetivo": (
            "Lançamento do perfil renovado + Prevenção cardiovascular\n"
            "Estreia com apresentação em SP (reposicionamento previsto para janeiro) ancorada no "
            "Dia Mundial da Saúde (7/abr). Estabelecer voz, identidade e localização."
        ),
        "posts": (
            "1. Apresentação — nova fase em SP (Conv)\n"
            "2. Checkup cardiovascular: quem precisa fazer? (MEV)\n"
            "3. MAPA: quando a pressão no consultório engana (MEV)\n"
            "4. Meus diferenciais: IC avançada e transplante (Conv)"
        ),
        "linhas":    "MEV (2) · Conv (2)",
        "datas":     "7/abr — Dia Mundial da Saúde",
        "funil":     "TOPO + FUNDO",
        "anuncio": (
            "Objetivo: reconhecimento de marca (lançamento)\n"
            "Impulsionar post 1 (Conv) + post 2 (MEV) por 7–10 dias cada.\n"
            "Público: mulheres e homens 40–70 anos, SP capital + raio 30 km,\n"
            "interesses: saúde, cardiologia, bem-estar.\n"
            "Orçamento sugerido: R$ 250–350 (lançamento)."
        ),
        "kpis": (
            "• Alcance e impressões\n"
            "• Novos seguidores\n"
            "• Visitas ao perfil\n"
            "• CPM (meta: < R$ 20)\n"
            "• Taxa de engajamento"
        ),
        "destaque": False,
    },
    {
        "mes":      "Maio",
        "num":      "2º mês",
        "objetivo": (
            "Saúde cardiovascular da mulher + Dia das Mães\n"
            "Usar o gancho emocional do Dia das Mães (10/mai) para conectar cuidado e saúde cardíaca. "
            "Incorporar temas de estresse e lifestyle previstos para fevereiro."
        ),
        "posts": (
            "1. Saúde cardiovascular da mulher (MEV)\n"
            "2. Arritmias e IC — sintomas a não ignorar (MEV)\n"
            "3. Cuidar da saúde = o melhor presente para quem você ama (Conv)\n"
            "4. Congresso / corrida com gancho cardiovascular (Hum) — se houver evento"
        ),
        "linhas":    "MEV (2) · Conv (1) · Hum (1)",
        "datas":     "10/mai — Dia das Mães",
        "funil":     "TOPO + FUNDO",
        "anuncio": (
            "Objetivo: alcance + tráfego\n"
            "Impulsionar post 3 (Conv — Dia das Mães) próximo a 10/mai.\n"
            "Público: mulheres 35–65, SP, + públicos semelhantes a seguidores atuais.\n"
            "Orçamento sugerido: R$ 150–200."
        ),
        "kpis": (
            "• Alcance orgânico vs pago\n"
            "• Cliques no link da bio\n"
            "• CTR (meta: > 1%)\n"
            "• Novos seguidores\n"
            "• Compartilhamentos"
        ),
        "destaque": False,
    },
    {
        "mes":      "Junho",
        "num":      "3º mês",
        "objetivo": (
            "Exercício físico e cardiologia\n"
            "Posicionar exercício como prescrição médica cardiovascular. "
            "Gancho do inverno (retomada de atividade). Apresentar VO2max e qualidade de vida pós-Tx."
        ),
        "posts": (
            "1. Por que cardiologista prescreve exercício? (MEV)\n"
            "2. VO2max: o número que prevê sua longevidade (MEV)\n"
            "3. Qualidade de vida pós-transplante (Causa)\n"
            "4. Onde atendo em SP e como agendar sua consulta (Conv)"
        ),
        "linhas":    "MEV (2) · Causa (1) · Conv (1)",
        "datas":     "—",
        "funil":     "TOPO + MEIO + FUNDO",
        "anuncio": (
            "Objetivo: tráfego / mensagens\n"
            "Impulsionar post 4 (Conv — como agendar) com objetivo 'Mensagens'.\n"
            "Público: pessoas que interagiram com o perfil nos últimos 60 dias +\n"
            "semelhantes a quem clicou no link.\n"
            "Orçamento sugerido: R$ 150–200."
        ),
        "kpis": (
            "• Mensagens iniciadas\n"
            "• CPC (meta: < R$ 3)\n"
            "• CPL — custo por lead\n"
            "• Cliques no link da bio\n"
            "• Salvamentos (qualidade)"
        ),
        "destaque": False,
    },
    {
        "mes":      "Julho",
        "num":      "4º mês",
        "objetivo": (
            "Posicionamento profissional + IC avançada\n"
            "Consolidar autoridade na especialidade de IC avançada e transplante. "
            "Aproveitar congresso para humanização com gancho médico."
        ),
        "posts": (
            "1. IC avançada: o Quarteto Fantástico (MEV)\n"
            "2. Colesterol e aterosclerose: a placa que se forma em silêncio (MEV)\n"
            "3. O que esperar da sua primeira consulta comigo (Conv)\n"
            "4. Congresso / bastidores do ambulatório pós-Tx (Hum) — se houver"
        ),
        "linhas":    "MEV (2) · Conv (1) · Hum (1)",
        "datas":     "Verificar: SOCESP / CBC",
        "funil":     "TOPO + MEIO + FUNDO",
        "anuncio": (
            "Objetivo: reconhecimento de marca (especialidade)\n"
            "Impulsionar post 1 (IC avançada) para alcançar pacientes com IC.\n"
            "Público: pessoas com interesse em 'insuficiência cardíaca', 'transplante',\n"
            "50–75 anos, SP.\n"
            "Orçamento sugerido: R$ 100–150."
        ),
        "kpis": (
            "• Alcance segmentado\n"
            "• CPM\n"
            "• Salvamentos\n"
            "• Visitas ao perfil\n"
            "• Novos seguidores qualificados"
        ),
        "destaque": False,
    },
    {
        "mes":      "Agosto",
        "num":      "5º mês",
        "objetivo": (
            "Fatores de risco cardiovascular + aquecimento para Setembro\n"
            "Resgatar temas de hipertensão, diabetes, sono e dieta que estavam previstos para fev/mar. "
            "Preparar audiência para a campanha de setembro."
        ),
        "posts": (
            "1. Dieta DASH: a mais recomendada por cardiologistas (MEV)\n"
            "2. Sono e o coração: 7 a 9 horas não é luxo (MEV)\n"
            "3. Fila de espera pelo transplante — dados reais (Causa)\n"
            "4. Por que consultar especialista em IC avançada? (Conv)"
        ),
        "linhas":    "MEV (2) · Causa (1) · Conv (1)",
        "datas":     "—",
        "funil":     "TOPO + MEIO + FUNDO",
        "anuncio": (
            "Objetivo: alcance + aquecimento de público\n"
            "Impulsionar post MEV de maior apelo (sono ou dieta) para ampliar audiência antes de set.\n"
            "Criar público personalizado de remarketing para usar em setembro.\n"
            "Orçamento sugerido: R$ 100–150."
        ),
        "kpis": (
            "• Alcance\n"
            "• Impressões\n"
            "• Taxa de engajamento\n"
            "• Crescimento de seguidores\n"
            "• Público de remarketing gerado"
        ),
        "destaque": False,
    },
    {
        "mes":      "Setembro ⭐",
        "num":      "6º mês",
        "objetivo": (
            "CAMPANHA: Mês do Coração — maior mês do ano\n"
            "Concentração máxima de causa e conscientização. Pico de alcance orgânico. "
            "5 posts (bônus de engajamento). Ancoragem nas datas mais importantes do calendário cardiovascular."
        ),
        "posts": (
            "1. Abertura Mês do Coração — dados de mortalidade cardiovascular (Causa)\n"
            "2. Como preservar o coração: guia prático (MEV)\n"
            "3. 27/set — Dia Nacional de Doação de Órgãos (Causa)\n"
            "4. 29/set — Dia Mundial do Coração: agende sua avaliação (Conv)\n"
            "BÔNUS: 'Você é doador?' — post de engajamento (Causa)"
        ),
        "linhas":    "Causa (3) · MEV (1) · Conv (1)",
        "datas":     "27/set — Dia Doação de Órgãos\n29/set — Dia Mundial do Coração",
        "funil":     "TOPO + FUNDO",
        "anuncio": (
            "INVESTIMENTO MÁXIMO DO ANO\n"
            "Impulsionar TODOS os posts da campanha. Estratégia dupla:\n"
            "• Posts de Causa (27/set e abertura): objetivo Alcance — público amplo SP.\n"
            "• Post Conv (29/set): objetivo Mensagens — público quente (remarketing de ago+set).\n"
            "Público frio: interesse em doação de órgãos, cardiologia, saúde — SP.\n"
            "Orçamento sugerido: R$ 400–600 (mês inteiro)."
        ),
        "kpis": (
            "• Alcance total da campanha\n"
            "• Impressões\n"
            "• CPM (eficiência)\n"
            "• Mensagens iniciadas\n"
            "• CPL — custo por lead\n"
            "• Crescimento de seguidores\n"
            "• Visitas ao perfil"
        ),
        "destaque": True,
    },
    {
        "mes":      "Outubro",
        "num":      "7º mês",
        "objetivo": (
            "Reconversão pós-campanha\n"
            "Aproveitar o aquecimento de setembro para converter novos seguidores em pacientes. "
            "Aprofundar temas técnicos e abrir espaço para Q&A."
        ),
        "posts": (
            "1. Miocardiopatias: quando o músculo do coração adoece (MEV)\n"
            "2. Medicamentos cardiológicos: dúvidas frequentes (MEV)\n"
            "3. Resultados do transplante cardíaco: dados reais (Causa)\n"
            "4. Q&A: respondendo suas dúvidas sobre o coração (Conv)"
        ),
        "linhas":    "MEV (2) · Causa (1) · Conv (1)",
        "datas":     "—",
        "funil":     "MEIO + FUNDO",
        "anuncio": (
            "Objetivo: conversão / mensagens\n"
            "Impulsionar post 4 (Q&A — Conv) com objetivo Mensagens.\n"
            "Público: remarketing de quem interagiu em setembro + semelhantes.\n"
            "Orçamento sugerido: R$ 150–200."
        ),
        "kpis": (
            "• Mensagens iniciadas\n"
            "• CPL\n"
            "• Taxa comentários (Q&A)\n"
            "• CPC\n"
            "• Conversão DM → consulta"
        ),
        "destaque": False,
    },
    {
        "mes":      "Novembro",
        "num":      "8º mês",
        "objetivo": (
            "Prevenção — saúde preventiva\n"
            "Mês de reforço preventivo. Tabagismo realocado de março. "
            "Humanização com aniversário ou congresso se houver. Reafirmar diferenciais antes do encerramento."
        ),
        "posts": (
            "1. Tabagismo e o coração (MEV)\n"
            "2. Teste ergométrico: por que o médico pediu a esteira? (MEV)\n"
            "3. Aniversário da Paôla ou congresso (Hum) — se aplicável\n"
            "4. Por que me escolher como sua cardiologista (Conv)"
        ),
        "linhas":    "MEV (2) · Hum (1) · Conv (1)",
        "datas":     "Verificar: Congresso Medicina do Estilo de Vida",
        "funil":     "TOPO + FUNDO",
        "anuncio": (
            "Objetivo: conversão\n"
            "Impulsionar post 4 (Conv — por que me escolher) com objetivo Mensagens.\n"
            "Público: seguidores há mais de 30 dias que ainda não interagiram com Conv.\n"
            "Orçamento sugerido: R$ 150–200."
        ),
        "kpis": (
            "• Mensagens iniciadas\n"
            "• CPC\n"
            "• CPL\n"
            "• Taxa de conversão DM → agendamento\n"
            "• Custo por consulta agendada"
        ),
        "destaque": False,
    },
    {
        "mes":      "Dezembro",
        "num":      "9º mês",
        "objetivo": (
            "Fechamento do ano + cuidados nas festas\n"
            "Encerrar 2026 com conteúdo prático para as festas, humanização (retrospectiva) "
            "e CTA de agendamento para 2027."
        ),
        "posts": (
            "1. Coração nas festas: álcool, sal e estresse (MEV)\n"
            "2. Coração artificial como ponte para o transplante (Causa)\n"
            "3. Retrospectiva 2026 (Hum)\n"
            "4. Agende seu checkup cardiovascular para 2027 (Conv)"
        ),
        "linhas":    "MEV (1) · Causa (1) · Hum (1) · Conv (1)",
        "datas":     "Festas de fim de ano",
        "funil":     "TOPO + FUNDO",
        "anuncio": (
            "Objetivo: conversão / agendamentos jan/2027\n"
            "Impulsionar post 4 (Conv — checkup 2027) com objetivo Mensagens na última semana de dez.\n"
            "Público: todas as pessoas que interagiram com o perfil em 2026.\n"
            "Orçamento sugerido: R$ 150–200."
        ),
        "kpis": (
            "• Mensagens iniciadas (meta: agendamentos para jan)\n"
            "• CPL\n"
            "• CPC\n"
            "• ROAS geral do ano (opcional)\n"
            "• Crescimento de seguidores 2026"
        ),
        "destaque": False,
    },
]

# ── Cabeçalho da aba Objetivos ────────────────────────────────────────────────
set_title(wo, get_column_letter(NCOLS), 1,
    "CALENDÁRIO EDITORIAL 2026 — OBJETIVOS & ESTRATÉGIA DE ANÚNCIOS MÊS A MÊS",
    HEADER_BG, "FFFFFF", size=13)

set_title(wo, get_column_letter(NCOLS), 2,
    "@paolapreto · Publicações a partir de ABRIL 2026 · "
    "9 meses ativos (abr–dez) · Temas de jan–mar redistribuídos em agosto e novembro",
    SUBHEADER_BG, "FFFFFF", size=10, bold=False, height=20)

wo.row_dimensions[3].height = 30
for ci, h in enumerate(headers_o, 1):
    c = wo.cell(row=3, column=ci, value=h)
    c.fill = fill(SUBHEADER_BG)
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    c.alignment = aln("center", wrap=True, v="center")
    c.border = Border(
        left=Side(style="medium", color="FFFFFF"),
        right=Side(style="medium", color="FFFFFF"),
        bottom=Side(style="medium", color="FFFFFF"),
    )

# ── Linhas de meses ───────────────────────────────────────────────────────────
for ri, m in enumerate(meses, start=4):
    wo.row_dimensions[ri].height = 120

    is_sept = m["destaque"]
    row_bg  = SEPT_BG if is_sept else (ROW_ALT if ri % 2 == 0 else ROW_BASE)
    ads_bg  = "E8F5E9" if not is_sept else "FFF9C4"

    vals = [
        m["mes"], m["num"], m["objetivo"], m["posts"],
        m["linhas"], m["datas"], m["funil"],
        m["anuncio"], m["kpis"],
    ]

    for ci, val in enumerate(vals, 1):
        c = wo.cell(row=ri, column=ci, value=val)
        c.border = border_light

        if ci == 1:    # Mês
            c.fill = fill(row_bg)
            c.font = Font(name="Calibri", color="1A3A4A" if not is_sept else "856404", bold=True, size=12)
            c.alignment = aln("center", v="center")
        elif ci == 2:  # Sequência
            c.fill = fill(row_bg)
            c.font = Font(name="Calibri", color="888888", size=9, italic=True)
            c.alignment = aln("center", v="center")
        elif ci == 3:  # Objetivo
            c.fill = fill(row_bg)
            c.font = fnt("2C3E50", size=10)
            c.alignment = aln("left", v="top")
        elif ci == 4:  # Posts
            c.fill = fill(row_bg)
            c.font = fnt("4A4A4A", size=9)
            c.alignment = aln("left", v="top")
        elif ci == 5:  # Linhas
            c.fill = fill(row_bg)
            c.font = fnt("1A5276", bold=True, size=9)
            c.alignment = aln("center", v="center")
        elif ci == 6:  # Datas
            c.fill = fill(row_bg)
            c.font = fnt("922B21" if m["datas"] not in ("—", "") else "AAAAAA",
                         bold=(m["datas"] not in ("—", "")), size=9)
            c.alignment = aln("center", v="center")
        elif ci == 7:  # Funil
            c.fill = fill(row_bg)
            c.font = fnt("1A3A4A", bold=True, size=9)
            c.alignment = aln("center", v="center")
        elif ci == 8:  # Anúncio
            c.fill = fill(ads_bg)
            c.font = fnt("1B5E20", size=9)
            c.alignment = aln("left", v="top")
        elif ci == 9:  # KPIs
            c.fill = fill(ads_bg)
            c.font = fnt("1B5E20", bold=True, size=9)
            c.alignment = aln("left", v="top")

    top_border(wo, ri, NCOLS)

# ── Seção: Guia de Campanhas de Anúncios ─────────────────────────────────────
gr = len(meses) + 6  # linha inicial da seção

# Título da seção
wo.merge_cells(f"A{gr}:{get_column_letter(NCOLS)}{gr}")
c = wo[f"A{gr}"]
c.value = "GUIA DE CAMPANHAS DE ANÚNCIOS — REFERÊNCIA"
c.fill = fill(ADS_BG)
c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=13)
c.alignment = aln("center", v="center")
wo.row_dimensions[gr].height = 32
gr += 1

# ── Sub-seção: Categorias de anúncio ──────────────────────────────────────────
wo.merge_cells(f"A{gr}:{get_column_letter(NCOLS)}{gr}")
c = wo[f"A{gr}"]
c.value = "CATEGORIAS DE ANÚNCIO"
c.fill = fill(ADS_SUB)
c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
c.alignment = aln("center", v="center")
wo.row_dimensions[gr].height = 24
gr += 1

cat_headers = ["CATEGORIA", "OBJETIVO", "QUANDO USAR", "CONTEÚDO IDEAL", "MÉTRICAS PRINCIPAIS", "ORÇAMENTO MENSAL"]
cat_widths  = [22, 28, 36, 30, 34, 22]
cat_col_map = [1, 2, 3, 5, 7, 9]  # colunas do grid de 9 cols

wo.row_dimensions[gr].height = 26
for col, h in zip(cat_col_map, cat_headers):
    c = wo.cell(row=gr, column=col, value=h)
    c.fill = fill("D4EDDA")
    c.font = Font(name="Calibri", color="155724", bold=True, size=9)
    c.alignment = aln("center", v="center")
    c.border = border_light
gr += 1

categorias = [
    (
        "Reconhecimento de Marca",
        "Alcançar novos públicos e aumentar seguidores qualificados",
        "Lançamento (abr), aquecimento (ago), Mês do Coração (set)",
        "Posts MEV de alto interesse geral (sono, exercício, dados de impacto)",
        "Alcance · Impressões · CPM · Novos seguidores · Visitas ao perfil",
        "R$ 100–200/mês",
    ),
    (
        "Engajamento",
        "Aumentar interações: curtidas, salvamentos, comentários",
        "Posts de séries (Mito ou Verdade, Q&A) com alto potencial viral",
        "Reels educativos, carrosséis informativos, Q&A",
        "Taxa de engajamento · Salvamentos · Comentários · Compartilhamentos",
        "R$ 80–120/mês",
    ),
    (
        "Tráfego",
        "Direcionar usuários para o link na bio (agendamento, site)",
        "Meses de posicionamento e apresentação de serviços",
        "Posts Conv com CTA claro para link na bio",
        "CTR · Cliques no link · CPC · Visitas ao site/WhatsApp",
        "R$ 100–150/mês",
    ),
    (
        "Mensagens (WhatsApp/DM)",
        "Gerar conversas diretas que convertam em consultas",
        "Meses de conversão: jun, out, nov, dez",
        "Posts Conv: onde atendo, primeira consulta, diferenciais",
        "Mensagens iniciadas · CPL · Taxa DM → agendamento · CPC",
        "R$ 150–200/mês",
    ),
    (
        "Campanha de Causa",
        "Ampliar alcance orgânico em datas simbólicas",
        "Setembro: Mês do Coração, 27/set Doação, 29/set Dia Mundial",
        "Posts de conscientização e doação de órgãos",
        "Alcance · Impressões · Compartilhamentos · CPM",
        "R$ 200–300 (set. apenas)",
    ),
]

for cat_row in categorias:
    wo.row_dimensions[gr].height = 50
    for col, val in zip(cat_col_map, cat_row):
        c = wo.cell(row=gr, column=col, value=val)
        c.fill = fill("F0FFF4")
        c.font = fnt("1B5E20", size=9)
        c.alignment = aln("left", v="top")
        c.border = border_light
    top_border(wo, gr, NCOLS)
    gr += 1

gr += 1  # espaço

# ── Sub-seção: KPIs de referência ─────────────────────────────────────────────
wo.merge_cells(f"A{gr}:{get_column_letter(NCOLS)}{gr}")
c = wo[f"A{gr}"]
c.value = "KPIs DE REFERÊNCIA — BENCHMARKS PARA INSTAGRAM ADS (SAÚDE, BRASIL)"
c.fill = fill(ADS_SUB)
c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
c.alignment = aln("center", v="center")
wo.row_dimensions[gr].height = 24
gr += 1

kpi_headers = ["KPI", "DEFINIÇÃO", "BENCHMARK REFERÊNCIA", "SINAL DE ALERTA", "AÇÃO CORRETIVA"]
kpi_col_map = [1, 2, 4, 6, 8]

wo.row_dimensions[gr].height = 26
for col, h in zip(kpi_col_map, kpi_headers):
    c = wo.cell(row=gr, column=col, value=h)
    c.fill = fill("D4EDDA")
    c.font = Font(name="Calibri", color="155724", bold=True, size=9)
    c.alignment = aln("center", v="center")
    c.border = border_light
gr += 1

kpis_ref = [
    ("CPM\n(Custo por Mil Impressões)",
     "Quanto custa alcançar 1.000 pessoas",
     "R$ 8–20 (saúde, SP)",
     "CPM > R$ 25",
     "Revisar segmentação de público; testar criativo diferente"),
    ("CTR\n(Taxa de Cliques no Link)",
     "% de pessoas que viram e clicaram no link/CTA",
     "1% – 3%",
     "CTR < 0,5%",
     "Melhorar o CTA; testar imagem ou texto de abertura diferentes"),
    ("CPC\n(Custo por Clique)",
     "Quanto custa cada clique no link ou botão",
     "R$ 1–4",
     "CPC > R$ 6",
     "Ajustar público; testar objetivo de campanha diferente"),
    ("CPL\n(Custo por Lead / Mensagem)",
     "Quanto custa cada mensagem ou contato iniciado",
     "R$ 15–50 (saúde especializada)",
     "CPL > R$ 70",
     "Revisar copy do post; testar público mais qualificado"),
    ("Taxa de Engajamento",
     "(Curtidas + Comentários + Salvamentos) ÷ Alcance",
     "2% – 6% (bom para perfil médico)",
     "< 1%",
     "Publicar em horários de pico; usar formatos de maior alcance (Reels)"),
    ("Crescimento de Seguidores",
     "Seguidores novos por mês",
     "+50–150/mês (fase inicial)",
     "< 20/mês ou perda",
     "Aumentar frequência de posts TOPO; revisar estratégia de anúncios de alcance"),
    ("Taxa DM → Consulta",
     "% de mensagens que se tornam agendamentos efetivos",
     "10% – 30% (estimativa)",
     "< 5%",
     "Revisar protocolo de resposta no DM; oferecer mais informações sobre a consulta"),
]

for kpi_row in kpis_ref:
    wo.row_dimensions[gr].height = 52
    for col, val in zip(kpi_col_map, kpi_row):
        c = wo.cell(row=gr, column=col, value=val)
        c.fill = fill("F0FFF4")
        c.font = fnt("1B5E20", size=9)
        c.alignment = aln("left", v="top")
        c.border = border_light
    top_border(wo, gr, NCOLS)
    gr += 1

gr += 1

# ── Sub-seção: Nota CFM ───────────────────────────────────────────────────────
wo.merge_cells(f"A{gr}:{get_column_letter(NCOLS)}{gr}")
c = wo[f"A{gr}"]
c.value = (
    "NOTA LEGAL (CFM — RESOLUÇÃO 2.336/2023):  "
    "Anúncios médicos no Instagram devem omitir preços, promoções, descontos e comparações com outros profissionais. "
    "Proibido o uso de depoimentos de pacientes como prova de eficácia. "
    "Imagens de antes/depois são vedadas. "
    "Todo anúncio deve identificar a profissional (nome + CRM). "
    "Conteúdos de causa e conscientização (setembro) têm menor risco regulatório — priorize-os para impulsionamento."
)
c.fill = fill("FFF3CD")
c.font = Font(name="Calibri", color="856404", size=9, italic=True)
c.alignment = aln("left", wrap=True, v="center")
wo.row_dimensions[gr].height = 52

# ── Congelar cabeçalho ────────────────────────────────────────────────────────
wo.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════════════
# SALVAR
# ══════════════════════════════════════════════════════════════════════════════
wb.save(XLSX)
print("Concluído. Abas geradas:")
for s in wb.sheetnames:
    print(f"  • {s}")
